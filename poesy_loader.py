import datetime
import time
import traceback

import requests
from bs4 import BeautifulSoup
import logging
from openpyxl import Workbook, load_workbook
import os
import asyncio
import aiohttp
import configparser

logger = logging.getLogger('poesy_loader')
logger.setLevel(logging.INFO)
handler = logging.FileHandler('run.log')
handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

MAX_TRY = 15


class PoesyLoader:

    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read('poesy_loader.ini')
        self.output_file_name = self.config['MAIN']['output_file']
        self.row = 1
        self.async_url_data = []
        self.url = self.config['MAIN']['url']
        self.page_count = self.get_page_count()
        self.poems_count = 0

    def get_poem_link(self, poem) -> dict:
        href = poem.find('a', class_='_2A3Np').get('href')
        link = '/'.join(self.url.split('/')[:-3]) + href
        return link

    @staticmethod
    def parse_poem(raw):
        resp = {}
        soup = BeautifulSoup(raw, 'lxml')

        div_raw = soup.find('div', class_='_1MTBU _3RpDE _47J4f _3IEeu')
        author = div_raw.find('div', class_='_14JnI').text
        name = div_raw.find('div', class_='_2jzeL').text
        strings_raw = div_raw.find('div', class_='_3P9bi')
        year_raw = strings_raw.find('div')
        year = ''
        if year_raw:
            year = f'\n\n{year_raw.text}'
        raw_p = strings_raw.find_all('p', class_='')
        quatrains = []
        for p in raw_p:
            quatrain = p.decode()
            quatrain = quatrain.replace('<p class="">', '')
            quatrain = quatrain.replace('</p>', '')
            quatrain = quatrain.replace('<br/>', '\n')
            quatrains.append(quatrain)
        poem = '\n\n'.join(quatrains)
        resp['author'] = author
        resp['name'] = name
        resp['poem'] = poem + year
        return resp

    def get_page_count(self) -> int:
        soup = BeautifulSoup(requests.get(self.url).text, 'lxml')
        logger.info(f'Get url: {self.url}')
        div_raw = soup.find('div', class_='_2uPBE')
        a_raw = div_raw.find_all('a', class_='GmJ5E')
        page_count = int(a_raw[-1].text)
        return page_count

    async def get_url(self, session, url) -> None:
        async with session.get(url) as res:
            data = await res.text()
            if res.status == 200:
                logger.info(f'Get successful ({url})')
            else:
                logger.error(f'Get unsuccessful ({url})')
            self.async_url_data.append(data)

    def get_page(self, page: int):
        current_try = 0
        while current_try < MAX_TRY:
            current_try += 1
            try:
                if page > 1:
                    url = self.url + f'?page={page}'
                else:
                    url = self.url
                res = requests.get(url).text
            except Exception:
                logger.exception(f'{traceback.format_exc()}')
                time.sleep(current_try)
            else:
                logger.info(f'Get url: {url}')
                return res
        logger.error(f'MAX_TRY exceeded')

    @staticmethod
    def get_poem_page(url):
        current_try = 0
        while current_try < MAX_TRY:
            current_try += 1
            try:
                text = requests.get(url).text
            except Exception:
                logger.exception(f'{traceback.format_exc()}')
                time.sleep(current_try)
            else:
                return text
        logger.error(f'MAX_TRY exceeded')

    def to_xls(self, array: list):
        '''
        Write array to xls
        :param array: input list of dict [{'author': 'name', 'name': 'name', 'poem': 'text'},
                                          {'author': 'name', 'name': 'name', 'poem': 'text'},
                                          ...,
                                          {'author': 'name', 'name': 'name', 'poem': 'text'}]
        '''
        # Проверка наличия файла и создание/присоединение, в зависимости от результата
        logger.info('Begin write to file')
        if not os.path.exists(self.output_file_name):
            logger.info(f'File {self.output_file_name} does not exist. Create file')
            wb = Workbook()
        else:
            logger.info(f'File {self.output_file_name} exist. Open file')
            wb = load_workbook(self.output_file_name)
        sheet = wb.active

        logger.info(f'Write into the file')
        for item in array:
            sheet['A' + str(self.row)] = item['author']
            sheet['B' + str(self.row)] = item['name']
            sheet['C' + str(self.row)] = item['poem']
            self.row += 1

        # Сохранить файл:
        wb.save(self.output_file_name)
        logger.info('End write to file')

    def main(self):
        now = datetime.datetime.now
        logger.info('<===================================================>')
        logger.info('PoesyLoader start')
        logger.info(f'Page count = {self.page_count}')
        for page in range(1, self.page_count + 1):
            logger.info(f'Get {page} page')
            start = now()
            text = self.get_page(page)
            logger.info(f'Parse {page} page')
            soup = BeautifulSoup(text, 'lxml')
            poems_raw = soup.find('div', class_='_2VELq')
            logger.info('Get list of poems')
            poems_raw = poems_raw.find_all('div', class_='_1jGw_')
            ready_to_write = []
            for num, poem_raw in enumerate(poems_raw):
                logger.info(f'Get {num+1} poem')
                link = self.get_poem_link(poem_raw)
                text = self.get_poem_page(link)
                logger.info(f'Parse {num+1} poem')
                ready_to_write.append(PoesyLoader.parse_poem(text))
                self.poems_count += 1
                logger.info(f'Complete {num+1} poem')
            self.to_xls(ready_to_write)
            finish = now() - start
            logger.info(f'Page {page} done. Time: {finish.seconds}.{str(finish.microseconds)[:3]} sec')
            logger.info(f'{self.poems_count} poems have been write into the file')
        logger.info('PoesyLoader finish\n')

    async def async_main(self):
        for page in range(1, self.page_count + 1):
            self.async_url_data = []
            async with aiohttp.ClientSession() as session:
                tasks = []
                if page > 1:
                    tasks.append(asyncio.create_task(self.get_url(session, self.url + f'?page={page}')))
                else:
                    tasks.append(asyncio.create_task(self.get_url(session, self.url)))
                await asyncio.gather(*tasks)
                soup = BeautifulSoup(self.async_url_data.pop(), 'lxml')
                poems_raw = soup.find('div', class_='_2VELq')
                poems_raw = poems_raw.find_all('div', class_='_1jGw_')
                tasks = []
                for poem_raw in poems_raw:
                    link = self.get_poem_link(poem_raw)
                    tasks.append(asyncio.create_task(self.get_url(session, link)))
                await asyncio.gather(*tasks)
                ready_to_write = []
                for raw in self.async_url_data:
                    ready_to_write.append(PoesyLoader.parse_poem(raw))
                self.to_xls(ready_to_write)
        logger.info('PoesyLoader finish\n')


if __name__ == "__main__":
    pl = PoesyLoader()
    pl.main()
    #asyncio.run(pl.main())
